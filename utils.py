from datetime import date
import os
import re
import time
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By

import openpyxl
import requests

from drivers import headless_driver


class DataToExcel:
    def __init__(self):
        self.cards_count = 0

    def create_excel(self):
        today = date.today().strftime("%d.%m.%y")
        path = os.path.join(os.path.dirname(__file__), f'./docs/companies_{today}.xlsx')
        if not os.path.exists(path):
            wb = openpyxl.Workbook()
            ws = wb.active

            ws.column_dimensions['A'].width = 27
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 40
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 25
            ws.column_dimensions['I'].width = 25
            ws.column_dimensions['J'].width = 20
            ws.column_dimensions['K'].width = 20
            ws.column_dimensions['L'].width = 40
            ws.column_dimensions['M'].width = 10
            ws.column_dimensions['N'].width = 20
            ws.column_dimensions['O'].width = 20
            ws.column_dimensions['P'].width = 20
            ws.column_dimensions['Q'].width = 15

            ws['A1'] = 'Сокращенное название'
            ws['B1'] = 'Полное название'
            ws['C1'] = 'Статус'
            ws['D1'] = 'ИНН'
            ws['E1'] = 'ОГРН'
            ws['F1'] = 'Юридический адрес'
            ws['G1'] = 'Город'
            ws['H1'] = 'Уставный капитал'
            ws['I1'] = 'Дата регистрации'
            ws['J1'] = 'ФИО Руководителя'
            ws['K1'] = 'Основной ОКВЭД'
            ws['L1'] = 'Основной ОКВЭД (название)'
            ws['M1'] = 'Количество сотрудников'
            ws['N1'] = 'Телефоны'
            ws['O1'] = 'Электронная почта'
            ws['P1'] = 'Сайты'
            ws['Q1'] = 'Доходы'
            
            for column in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=column)
                cell.font = openpyxl.styles.Font(bold=True)

            wb.save(f'./docs/companies_{today}.xlsx')

    def collect_data(self, text, proxy):
        short_name = re.search(r'(.+)\s-', text).group(1)
        print(short_name)
        full_name = re.search(r'\n(.+)\s+ИНН', text).group(1)
        print(full_name)
        status = re.search(r'-\s(.+)\n', text).group(1)
        print(status)
        inn = re.search(r'ИНН\s(\d+),', text).group(1)
        print(inn)
        ogrn = re.search(r'ОГРН\s(\d+),', text).group(1)
        print(ogrn)
        address = re.search(r"\s+(\d{6},.+|.+)\s+ОК", text)
        address = '' if address is None else address.group(1)
        print(address)
        city = re.search(r'\s\d{6}.+(г\.\w+),', text)
        city = address.split(',')[1] if city is None else city.group(1)
        print(city)
        authorized_capital = re.search(r'капитал\s(.+)руб.,', text)
        authorized_capital = '' if authorized_capital is None else authorized_capital.group(1)
        print(authorized_capital)
        registered_date = re.search(r'(\d{4}-\d{2}-\d{2})', text)
        registered_date = '' if registered_date is None else registered_date.group(1)
        print(registered_date)

        manager_name = re.search(r'[Дд]иректор\s(.+)\n', text)
        manager_name = '' if manager_name is None else manager_name.group(1)
        print(manager_name)
        revenues = re.search(r'доходы\s(.+)руб', text)
        revenues = '' if revenues is None else revenues.group(1)
        print(revenues)

        code = re.search(r'"ОКВЭД\s(\S{2,})\s"', text)
        code = '' if code is None else code.group(1)
        
        activity_name = re.search(r"ОКВЭД\s\S{2,}\s(.+)", text)
        activity_name = '' if activity_name is None else activity_name.group(1)
        print(activity_name)

        employees = parse_sbis(inn, ogrn)
        phone, email, site = parse_contacts(ogrn, proxy)

        company_data = [
            [short_name, full_name, status, inn, ogrn, address, city, authorized_capital, registered_date, manager_name,
             code, activity_name, employees, phone, email, site, revenues]]

        self.add_data_to_excel(company_data)

    def add_data_to_excel(self, data):
        today = date.today().strftime("%d.%m.%y")
        workbook = openpyxl.load_workbook(f'./docs/companies_{today}.xlsx')
        sheet = workbook.active

        for row in data:
            sheet.append(row)

        workbook.save(f'./docs/companies_{today}.xlsx')
        print(f"Данные успешно добавлены в Excel файл: companies_{today}.xlsx, всего {self.cards_count} записей")
        self.cards_count += 1




def parse_contacts(ogrn, proxy):
    driver = headless_driver(proxy)
    driver.get(f'https://companium.ru/id/{ogrn}/contacts')
    time.sleep(2)
    phone = driver.find_elements(By.XPATH, '//a[contains(@id, "copy-phone")]')
    phones = []
    if len(phone) > 1:
        for i in phone:
            phones.append(i.text)
        phone = ', '.join(phones)
    elif len(phone) == 1:
        phone = phone[0].text
    else:
        phone = ''
    soup = BeautifulSoup(driver.page_source, 'lxml')
    email = soup.find('div', attrs={'class': 'col-12 col-lg-4 border-lg-start'}).find_all('a')
    emails = []
    sites = []
    for i in email:
        if '@' in i.text :
            emails.append(i.text)
            print(i.text)
        else:
            sites.append(i.text)
    if len(emails) > 1:
        email = ', '.join(emails)
    elif len(emails) == 1:
        email = emails[0]
    else:
        email = ''
    if len(sites) > 1:
        site = ', '.join(sites)
    elif len(sites) == 1:
        site = sites[0]
    else:
        site = ''
    return phone, email, site
    
    
def parse_sbis(inn, ogrn):
    cookies = {
    'lang': 'ru',
    's3reg': '63',
    'tmr_lvid': '12a5f101d52cc36ee43202d87ba1fd24',
    'tmr_lvidTS': '1693386726080',
    '_ym_uid': '1693386726562402242',
    '_ym_d': '1693386726',
    'DeviceId': 'ec011b4d-53a4-4e08-9e16-9302908dfd3d',
    's3sid-online-daab': '00a28c6f-00a2a4ae-000d-1111111111111111',
    '_ym_isad': '1',
    'header_phone': '%7B%22main%22:null,%22ad%22:null,%22time%22:null%7D',
    '_ym_visorc': 'b',
    'tmr_detect': '1%7C1695569105847',
    'tz': '-240',
    's3ds': '1920%7C1080%7C1210%7C976%7C1920%7C1053',
    'adaptiveAspects': '%7B%22windowInnerWidth%22%3A1210%2C%22windowInnerHeight%22%3A976%2C%22containerClientWidth%22%3A1210%2C%22containerClientHeight%22%3A976%2C%22isVertical%22%3Afalse%2C%22isTouch%22%3Afalse%7D',
}
    headers = {
    'authority': 'sbis.ru',
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'accept-language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
    'cache-control': 'max-age=0',
    # 'cookie': 'lang=ru; s3reg=63; tmr_lvid=12a5f101d52cc36ee43202d87ba1fd24; tmr_lvidTS=1693386726080; _ym_uid=1693386726562402242; _ym_d=1693386726; DeviceId=ec011b4d-53a4-4e08-9e16-9302908dfd3d; s3sid-online-daab=00a28c6f-00a2a4ae-000d-1111111111111111; _ym_isad=1; header_phone=%7B%22main%22:null,%22ad%22:null,%22time%22:null%7D; _ym_visorc=b; tmr_detect=1%7C1695569105847; tz=-240; s3ds=1920%7C1080%7C1210%7C976%7C1920%7C1053; adaptiveAspects=%7B%22windowInnerWidth%22%3A1210%2C%22windowInnerHeight%22%3A976%2C%22containerClientWidth%22%3A1210%2C%22containerClientHeight%22%3A976%2C%22isVertical%22%3Afalse%2C%22isTouch%22%3Afalse%7D',
    'sec-ch-ua': '"Google Chrome";v="117", "Not;A=Brand";v="8", "Chromium";v="117"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Linux"',
    'sec-fetch-dest': 'document',
    'sec-fetch-mode': 'navigate',
    'sec-fetch-site': 'same-origin',
    'sec-fetch-user': '?1',
    'upgrade-insecure-requests': '1',
    'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36',
}
    response = requests.get(f'https://sbis.ru/contragents/{inn}/{ogrn}', cookies=cookies, headers=headers)
    soup = BeautifulSoup(response.text, 'lxml')
    
    employees = soup.find('span', attrs={'itemprop':'numberOfEmployees'})
    if employees:
        employees = employees.text
    
    return employees


data_to_excel = DataToExcel()
